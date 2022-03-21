from asyncio.windows_events import NULL
from Public import *
from ast import If, Return
import xlrd
import sys
import os
import openpyxl
# import tkinter as tk
# from tkinter import filedialog
import matplotlib.pyplot as plt
from matplotlib.animation import FuncAnimation
import seaborn as sns
import numpy as np
import copy
import math

# CURRENT_PATN = os.path.dirname(os.path.realpath(sys.argv[0]))

TIMEDIFFCOL = 33
# WAVETIMECOL = 41
WAVETIMECOL = 55
CACHESIZE = 24  
MAXABNORMAL = 6

def next_index(index):
    if index < 23:
        index += 1
    else:
        index = 0
    return index

def before_index(index):
    if index > 0:
        index -= 1
    else:
        index = 23
    return index

def CalDistance(back,forward):  
    return (back + CACHESIZE - forward)%CACHESIZE
    
def CalTimeDiffus(origin):
    result = ((((origin%65536)/3)/65535)+((origin/65536)/3))/4
    return round( result,5)

def CalFlowm3h(timediff):
    result = CalTimeDiffus(timediff) *6.41817998
    return round( result,5)

def CalSumFlowL(timediff):
    result = CalTimeDiffus(timediff) *6.41817998/3.6
    return round( result,5)

def CalAvaerageFlowRate(timedifflist):
    max_value = max(timedifflist) 
    min_value = min(timedifflist)
    sum_value = sum(timedifflist)

    return int( (sum_value-max_value-min_value)/(len(timedifflist)-2))

def average_TOF(list,flag,star_index,end_index,num):
    average = 0
    
    index = star_index
    abnormal_sum = list[star_index]
    if flag == 1:
        average = np.mean(list)
    elif flag == 2:
        if num > 1:
            for i in range(24):
                index = next_index(index)
                abnormal_sum += list[index]
                if  index == end_index:
                    break
        average = (np.sum(list) - abnormal_sum)/(24-num)
    else:
        average = 0
    return average

# figm, axm = plt.subplots(figsize=(24, 8))  
# axm2 = axm.twinx()
# axm_e = axm.twinx()
# axm2_e = axm.twinx()
# def update(frame_data_sub):
#     total_time_diff_x = np.zeros(24)
#     total_time_diff = np.zeros(24)
#     total_wave_time = np.zeros(24)

#     length = len(frame_data_sub[1][0])
#     if length > 0:
#         total_x_axis_e = np.zeros(length)
#         total_time_diff_e = np.zeros(length)
#         total_wave_time_e = np.zeros(length)
#         for i in range(length):
#             total_x_axis_e[i] = frame_data_sub[1][0][i] 
#             total_wave_time[i] = frame_data_sub[1][1][i] 
#             total_time_diff_e[i] = frame_data_sub[1][2][i] 


#     for i in range(24):
#         total_time_diff_x[i] = frame_data_sub[0][i]
#         total_time_diff[i] = frame_data_sub[0][i+24]
#         total_wave_time[i] = frame_data_sub[0][i+48]
#         # total_time_diff_e[i] = frame_data_sub[i+72]
#         # total_wave_time_e[i] = frame_data_sub[i+96]



#     min_time_diff_x = np.min(total_time_diff_x)
#     max_time_diff_x = np.max(total_time_diff_x)

#     min_time_diff = np.min(total_time_diff)-1000
#     max_time_diff = np.max(total_time_diff)+1000

#     axm.clear()
#     axm.set_xlim(min_time_diff_x,max_time_diff_x)
#     axm.set_ylim(min_time_diff,max_time_diff)
#     axm.plot(total_time_diff_x, total_time_diff, "b", linewidth=1,marker='.')    # 加载曲线

#     for i in range(24):
#         axm.text(total_time_diff_x[i],total_time_diff[i],str(total_time_diff[i]))

#     axm.set_ylabel('time_diff')
#     axm.set_title('time diff distribution')

#     axm2.clear()
#     axm2.set_ylabel('wave_time')
#     min_wave_time = np.min(total_wave_time)-100
#     max_wave_time = np.max(total_wave_time)+100
#     axm2.set_xlim(min_time_diff_x,max_time_diff_x)
#     axm2.set_ylim(min_wave_time,max_wave_time)
#     axm2.plot(total_time_diff_x, total_wave_time, "r", linewidth=1,marker='.')    # 加载曲线
#     for i in range(24):
#         axm2.text(total_time_diff_x[i],total_wave_time[i],str(total_wave_time[i]))

#     axm_e.clear()

#     if length > 0:
#         axm_e.plot(total_x_axis_e, total_time_diff_e, "g", linewidth=1,marker='.')    # 加载曲线
#         for i in range(length):
#             axm_e.text(total_x_axis_e[i],total_time_diff_e[i],str(total_time_diff_e[i]))

#     axm2_e.clear()

   
#     if length > 0:
#         axm2_e.plot(total_x_axis_e, total_wave_time_e, 'purple', linewidth=1,marker='.')    # 加载曲线
#         for i in range(length):
#             axm2_e.text(total_x_axis_e[i],total_wave_time_e[i],str(total_wave_time_e[i]))

def LoadFlowData(self):
    # root = tk.Tk()
    # root.withdraw()

    # file_path = filedialog.askopenfilename() # 
    initialdir=os.path.dirname(os.path.realpath(sys.argv[0]))
    FileDialog = QFileDialog()
    # file_path, selectedFilter = FileDialog.getOpenFileName(tr("Open File"),()
    #                                             os.path.dirname(os.path.realpath(sys.argv[0])),
    #                                             tr("Excel (*.xls *.xlxs)"))

    file_path, selectedFilter = FileDialog.getOpenFileName(None,"OpenFile",initialdir,'Excel(*.xls *.xlsx)')
    book = xlrd.open_workbook(file_path)
    sheet = book.sheet_by_index(0)


    if book.nsheets > 1:
        sheet2 = book.sheet_by_index(1)
        man_flag  = 1
    else:
        man_flag  = 0

    fv.OriginFlowData.clear()
    fv.AutoExcFlowData.clear()
    fv.ManualExcFlowData.clear()

    datanrows = sheet.nrows

    realTm = {'wave_time' : [],'time_diff' : [],'man_time_diff' : [],'select_wave_time' : [],'select_time_diff' : [],'abnormal_wave_time_index' : [],'abnormal_time_diff_index' : [],'x_axis':[],'average_wave_time' : 0,'average_time_diff' : 0,'max_wave_time' : 0,'min_wave_time' : 0,'max_time_diff' : 0,'min_time_diff' : 0,'range' : 0,'record_cnt' : 0,'flag' : 0}
    realTm['wave_time'] =  [0 for x in range(0, 24)]
    realTm['time_diff'] =  [0 for x in range(0, 24)]
    realTm['man_time_diff'] =  [0 for x in range(0, 24)]
    realTm['select_wave_time'] =  [0 for x in range(0, 24)]
    realTm['select_time_diff'] =  [0 for x in range(0, 24)]
    realTm['x_axis'] =  [0 for x in range(0, 24)]
    realTm['abnormal_wave_time_index'] =  [0 for x in range(0, MAXABNORMAL)]
    realTm['abnormal_time_diff_index'] =  [0 for x in range(0, MAXABNORMAL)]

    hisTm =  {'wave_time' : [],'time_diff' : [],'x_axis':[]}
    curTm =  {'wave_time' : [],'time_diff' : [],'man_time_diff' : [],'x_axis':[]}

    sumflow = 0
    sumflowsub = 0
    flowrate = 0
    man_flowrate = 0
    origin_flowrate = 0
    flow = 0
    man_flow = 0
    origin_flow = 0
    flow_diff = 0
    om_flow_diff = 0
    aver_auto_flow = 0
    aver_man_flow = 0
    aver_origin_flow = 0

    alarmsumflowsub = 0
    alarmsumflow = 0

    auto_flow_cache = [0 for x in range(0, 6)]
    man_flow_cache = [0 for x in range(0, 6)]
    origin_flow_cache = [0 for x in range(0, 6)]
    cache_index = 0

    result = openpyxl.Workbook()
    total = result.active
    total.title = '汇总'

    total['A1'] = '管程时间'
    total['B1'] = '管程时间'
    total['C1'] = '管程时间'
    total['D1'] = '管程时间'
    total['E1'] = '管程时间'
    total['F1'] = '管程时间'
    total['G1'] = '管程时间'
    total['H1'] = '管程时间'
    total['I1'] = '时差'
    total['J1'] = '时差'
    total['K1'] = '时差'
    total['L1'] = '时差'
    total['M1'] = '时差'
    total['N1'] = '时差'
    total['O1'] = '时差'
    total['P1'] = '时差'
    total['Q1'] = '累计流量'
    total['R1'] = '瞬时流量'
    total['S1'] = '流量状态'
    total['T1'] = '时差状态'
    total['U1'] = '时差异常索引'
    total['V1'] = '管程状态'
    total['W1'] = '管程异常索引'


    total_time_diff = np.zeros(8*(datanrows-1))
    total_wave_time = np.zeros(8*(datanrows-1))
    total_time_diff_x = np.zeros(8*(datanrows-1))
    total_wave_time_x = np.zeros(8*(datanrows-1))

    for meter_date_rows in range(1, datanrows-2):
        timediff = np.zeros(24) 
        wavetime = np.zeros(24) 
        for temp_rows in range(meter_date_rows, meter_date_rows+3):
            for temp_cols in range(TIMEDIFFCOL, TIMEDIFFCOL+8):
                timediff[(temp_rows-meter_date_rows)*8+(temp_cols-TIMEDIFFCOL)] = sheet.cell(temp_rows,temp_cols).value
        for temp_rows in range(meter_date_rows, meter_date_rows+3):
            for temp_cols in range(WAVETIMECOL, WAVETIMECOL+8):
                wavetime[(temp_rows-meter_date_rows)*8+(temp_cols-WAVETIMECOL)] = sheet.cell(temp_rows,temp_cols).value
                
    for meter_date_rows in range(1, datanrows):  
        for temp_cols in range(TIMEDIFFCOL, TIMEDIFFCOL+8):
            total_time_diff[(meter_date_rows-1)*8+(temp_cols-TIMEDIFFCOL)] = sheet.cell(meter_date_rows,temp_cols).value
            total_time_diff_x[(meter_date_rows-1)*8+(temp_cols-TIMEDIFFCOL)] = (meter_date_rows - 1) + (temp_cols-TIMEDIFFCOL)*0.125
        for temp_cols in range(WAVETIMECOL, WAVETIMECOL+8):
            total_wave_time[(meter_date_rows-1)*8+(temp_cols-WAVETIMECOL)] = sheet.cell(meter_date_rows,temp_cols).value
            total_wave_time_x[(meter_date_rows-1)*8+(temp_cols-WAVETIMECOL)] = (meter_date_rows - 1) + (temp_cols-WAVETIMECOL)*0.125
    frame_data = []

    for i in range(datanrows-3):
        frame_data.append([])
        frame_data[i].append([])
        frame_data[i][0] = [0 for i in range(72)]
        frame_data[i].append([])
        frame_data[i][1].append([])
        frame_data[i][1].append([])
        frame_data[i][1].append([])
        for temp_rows in range(3):
            for temp_cols in range(8):
                frame_data[i][0][temp_rows*8+temp_cols] = total_time_diff_x[(i+temp_rows)*8+temp_cols]
            for temp_cols in range(8):
                frame_data[i][0][temp_rows*8+temp_cols+24] = total_time_diff[(i+temp_rows)*8+temp_cols]
            for temp_cols in range(8):
                frame_data[i][0][temp_rows*8+temp_cols+48] = total_wave_time[(i+temp_rows)*8+temp_cols]

    # ani = FuncAnimation(figm, update, frame_data, interval=1000.0/12, save_count=100)

    # print('Begin saving gif')
    # result_file_name = file_path.split('.', 1)
    # result_file_name = ''.join(result_file_name[0])
    # result_file_name = result_file_name + '.gif'
    # ani.save(result_file_name, writer='imagemagick', fps=0.5)
    # print('Finished.')


    his_flow = 0
    his_flow_flag = 0
    frame_data_e = []
    cal_full_flag = 0
    legal_cnt = 0
    record_index = 0
    record_cnt = 0
    his_cal_flag = 0
    his_flowrate_flag = 0
    his_flowrate2 = 0
    his_flowrate_flag2 = 0
    flowdic= {'flow_x':[0,1,2],'flow_rate':[0,0,0],'man_flow_rate':[0,0,0],'origin_flow_rate':[0,0,0],'flow':[0,0,0],'man_flow':[0,0,0],'flow_diff':[0,0,0],'origin_flow':[0,0,0],'om_flow_diff':[0,0,0]}
    for row in range(1,datanrows):
        alarmflag = 0
        
        cal_flag = 0
        legal_flag = 0


        for i in range(0,8):
            total.cell(row+2, i+1).value = sheet.cell_value (rowx=row, colx=i)
            total.cell(row+2, i+9).value = sheet.cell_value (rowx=row, colx=i+8)

            if sheet.cell_value (rowx=row, colx=i+WAVETIMECOL) != 0:
                realTm['time_diff'][record_index] = sheet.cell_value (rowx=row, colx=i+TIMEDIFFCOL)
                if man_flag == 1:
                    realTm['man_time_diff'][record_index] = sheet2.cell_value (rowx=row, colx=i+TIMEDIFFCOL)
                realTm['wave_time'][record_index] = sheet.cell_value (rowx=row, colx=i+WAVETIMECOL)
                realTm['x_axis'][record_index] = ((row-1)+0.125*i)
                if record_index < 23:
                    record_index += 1
                else:
                    record_index = 0

                if record_cnt < 23:
                    record_cnt += 1
                else: 
                    cal_full_flag = 3 

                legal_flag = 1
                legal_cnt = 3
            
            

                


     

        if legal_flag == 1:
            
            if cal_full_flag == 3:
                select_wave_time_flag = 0
                select_wave_time_cnt = 0
                select_wave_time_index = 0
                select_time_diff_flag = 0
                select_time_diff_cnt = 0
                select_time_diff_index = 0

                skip_index_start = 0
                skip_index_end = 0
                skip_index_distance = 0

                judge_index = record_index
                for i in range(0,23): 
                    wave_diff = realTm['wave_time'][next_index(judge_index)] - realTm['wave_time'][judge_index] 
                    judge_index = next_index(judge_index)
                    if wave_diff > 24:
                        realTm['select_wave_time'][judge_index] = 1
                        if select_wave_time_cnt < MAXABNORMAL:
                            realTm['abnormal_wave_time_index'][select_wave_time_cnt] = judge_index
                        select_wave_time_cnt += 1
                        if select_wave_time_cnt == 1:
                            select_wave_time_index = judge_index
                    elif wave_diff < -24:
                        realTm['select_wave_time'][judge_index] = -1
                        if select_wave_time_cnt < MAXABNORMAL:
                            realTm['abnormal_wave_time_index'][select_wave_time_cnt] = judge_index
                        select_wave_time_cnt += 1
                        if select_wave_time_cnt == 1:
                            select_wave_time_index = judge_index
                    else:
                        realTm['select_wave_time'][judge_index] = 0
                
                if select_wave_time_cnt > MAXABNORMAL:
                    select_wave_time_flag = 4
                elif select_wave_time_cnt >= 2:
                    abnormal_distance =  CalDistance(realTm['abnormal_wave_time_index'][select_wave_time_cnt-1],realTm['abnormal_wave_time_index'][0])
                    if abnormal_distance ==  1: #奇异点相邻
                        if realTm['select_wave_time'][realTm['abnormal_wave_time_index'][0]] + realTm['select_wave_time'][realTm['abnormal_wave_time_index'][1]] == 0:
                            select_wave_time_flag = 2
                        else:
                            select_wave_time_flag = 3
                    elif abnormal_distance < 6: #间距小于6
                        select_wave_time_flag = 3
                    else:
                        select_wave_time_flag = 4
                elif select_wave_time_cnt == 1:
                    select_wave_time_flag = 1
                else:
                    select_wave_time_flag = 0

                if select_wave_time_flag <= 3:
                    skip_index_start = 0
                    skip_index_end = 0
                    skip_index_distance = 0
                    if select_wave_time_flag == 2:
                        skip_index_start = realTm['abnormal_wave_time_index'][0]
                        skip_index_end = realTm['abnormal_wave_time_index'][0]
                        skip_index_distance = 1
                    elif select_wave_time_flag == 3:
                        skip_index_start = before_index(realTm['abnormal_wave_time_index'][0])
                        skip_index_end = next_index(realTm['abnormal_wave_time_index'][select_wave_time_cnt-1])
                        skip_index_distance = CalDistance(skip_index_end,skip_index_start) + 1
                    judge_index = record_index
                    if select_wave_time_flag > 1:
                        temprange = ((skip_index_end + CACHESIZE - skip_index_start)%CACHESIZE) - ((judge_index + CACHESIZE - skip_index_start)%CACHESIZE)
                        if temprange >=0: #起始地址在跳过区域
                            judge_index = next_index(skip_index_end)
                    for i in range(0,23-skip_index_distance):  
                        temp_next_index = next_index(judge_index)
                        if select_wave_time_flag > 1:
                            temprange = ((skip_index_end + CACHESIZE - skip_index_start)%CACHESIZE) - ((temp_next_index + CACHESIZE - skip_index_start)%CACHESIZE)
                            if temprange >=0: #地址在跳过区域
                                temp_next_index = next_index(skip_index_end)
                        time_diff = realTm['time_diff'][temp_next_index] - realTm['time_diff'][judge_index] 
                        judge_index = temp_next_index

                        if time_diff > 600000:
                            realTm['select_time_diff'][judge_index] = 1
                            select_time_diff_cnt += 1
                            if select_time_diff_cnt == 1:
                                select_time_diff_index = judge_index
                        elif time_diff < -600000:
                            realTm['select_time_diff'][judge_index] = -1
                            select_time_diff_cnt += 1
                            if select_time_diff_cnt == 1:
                                select_time_diff_index = judge_index
                        else:
                            realTm['select_time_diff'][judge_index] = 0
                    
                    if select_time_diff_cnt > 1:
                        select_time_diff_flag = 2
                    # elif select_time_diff_cnt == 2:
                    #     if realTm['select_time_diff'][select_time_diff_index] + realTm['select_time_diff'][next_index(select_time_diff_index)] != 0:
                    #         select_time_diff_flag = 3
                    #     else:
                    #         select_time_diff_flag = 1
                    elif select_time_diff_cnt == 1:
                        select_time_diff_flag = 1
                    else:
                        select_time_diff_flag = 0

                    if select_wave_time_flag == 0:
                        if select_time_diff_flag == 0:
                            cal_flag = 1
                        elif select_time_diff_flag == 1:
                            cal_flag = 2
                        else:
                            cal_flag = 3
                    elif select_wave_time_flag == 1:
                        if select_time_diff_flag == 0:
                            cal_flag = 1
                        elif select_time_diff_flag == 1 and select_time_diff_index == select_wave_time_flag:
                            cal_flag = 2
                        else:
                            cal_flag = 3
                    elif select_wave_time_flag == 2:
                        if select_time_diff_flag == 0:
                            cal_flag = 2
                        elif select_time_diff_flag == 1 and select_time_diff_index == select_wave_time_flag:
                            cal_flag = 2
                        else:
                            cal_flag = 3
                    elif select_wave_time_flag == 3:
                        if select_time_diff_flag <= 1:
                            cal_flag = 2
                        # elif select_time_diff_flag == 1:
                        #     cal_flag = 2
                        else:
                            cal_flag = 3
                else:
                    cal_flag = 3
                    
                # if cal_flag == 2:
                #     tof_average = (np.sum(realTm['time_diff']) - realTm['time_diff'][select_time_diff_index])/23
                #     realTm['time_diff'][select_time_diff_index] = tof_average

                flowrate = average_TOF(realTm['time_diff'],cal_flag,skip_index_start,skip_index_end,skip_index_distance)
                man_flowrate = average_TOF(realTm['man_time_diff'],1,skip_index_start,skip_index_end,skip_index_distance)
                origin_flowrate = average_TOF(realTm['time_diff'],1,skip_index_start,skip_index_end,skip_index_distance)
                if cal_flag == 1:
                    his_flowrate = flowrate
                    his_flowrate_flag = fv.NormalhisSize

                    his_flowrate2 = 0
                    his_flowrate_flag2 = 0
                    judge_index = record_index
                    for i in range(0,24):   
                        hisTm['time_diff'].append(CalTimeDiffus(realTm['time_diff'][judge_index]))
                        hisTm['wave_time'].append(realTm['wave_time'][judge_index])
                        hisTm['x_axis'].append(realTm['x_axis'][judge_index])
                        judge_index = next_index(judge_index)
                else:
                    if his_flowrate_flag > 0:
                        his_flowrate_flag -= 1
                    if  cal_flag == 2:
                        his_flowrate2 = flowrate
                        his_flowrate_flag2 = fv.AbnorhisSize
                        if his_flowrate_flag < fv.AbnorhisSize:
                            his_flowrate_flag = 0
                            his_flowrate = 0
                        judge_index = record_index
                        for i in range(0,24):   
                            temprange = ((skip_index_end + CACHESIZE - skip_index_start)%CACHESIZE) - ((judge_index + CACHESIZE - skip_index_start)%CACHESIZE)
                            if temprange >=0: #起始地址在跳过区域
                                judge_index = next_index(judge_index)
                                continue
                            hisTm['time_diff'].append(CalTimeDiffus(realTm['time_diff'][judge_index]))
                            hisTm['wave_time'].append(realTm['wave_time'][judge_index])
                            hisTm['x_axis'].append(realTm['x_axis'][judge_index])
                            judge_index = next_index(judge_index)
                    elif cal_flag == 3:
                        if his_flowrate_flag > 0:
                            flowrate =  his_flowrate
                        elif his_flowrate_flag2 > 0:
                            flowrate =  his_flowrate2
                            his_flowrate_flag2 -= 1
                        else:
                            flowrate = 0
                        judge_index = record_index
                        for i in range(0,24):   
                            hisTm['time_diff'].append(CalTimeDiffus(realTm['time_diff'][judge_index])) 
                            hisTm['wave_time'].append(realTm['wave_time'][judge_index])
                            hisTm['x_axis'].append(realTm['x_axis'][judge_index])
                            judge_index = next_index(judge_index)
                judge_index = record_index
                for i in range(0,24):   
                    curTm['time_diff'].append(CalTimeDiffus(realTm['time_diff'][judge_index])) 
                    curTm['man_time_diff'].append(CalTimeDiffus(realTm['man_time_diff'][judge_index])) 
                    curTm['wave_time'].append(int(realTm['wave_time'][judge_index]))
                    curTm['x_axis'].append(realTm['x_axis'][judge_index])
                    judge_index = next_index(judge_index)

                
                
                # if cal_flag > 1 and his_flowrate_flag > 0:
                #     if select_wave_time_flag > 0:
                #         record_index = select_wave_time_index
                #     elif select_time_diff_flag > 0:  
                #         record_index = select_time_diff_index 
                # his_cal_flag = cal_flag
        else:
            if legal_cnt > 0:
                legal_cnt -= 1
            else :
                cal_full_flag = 0
                record_cnt = 0
                realTm['wave_time'] =  [0 for x in range(0, 24)]
                realTm['time_diff'] =  [0 for x in range(0, 24)]
                realTm['select_wave_time'] =  [0 for x in range(0, 24)]
                realTm['select_time_diff'] =  [0 for x in range(0, 24)]

        if alarmflag == 1:
            alarmsumflowsub += flowrate
        else:
            sumflowsub += flowrate
        
        if alarmsumflowsub > 7200000:
            alarmsumflowsub = 0
        elif alarmsumflowsub > 3600000:
            alarmsumflow += 1
            alarmsumflowsub -= 3600000

        if sumflowsub > 7200000:
            sumflowsub = 0
        elif sumflowsub > 3600000:
            sumflow += 1
            sumflowsub -= 3600000

        auto_flow_cache.append(flowrate)
        del(auto_flow_cache[0])
        man_flow_cache.append(man_flowrate)
        del(man_flow_cache[0])
        origin_flow_cache.append(origin_flowrate)
        del(origin_flow_cache[0])

        aver_auto_flow = CalAvaerageFlowRate(auto_flow_cache)
        aver_man_flow = CalAvaerageFlowRate(man_flow_cache)
        aver_origin_flow = CalAvaerageFlowRate(origin_flow_cache)

        # flow += CalSumFlowL(flowrate) 
        # man_flow += CalSumFlowL(man_flowrate)
        # flow_diff = flow - man_flow

        flow += CalSumFlowL(aver_auto_flow) 
        man_flow += CalSumFlowL(aver_man_flow)
        origin_flow += CalSumFlowL(aver_origin_flow)
        flow_diff = flow - man_flow
        om_flow_diff = origin_flow - man_flow


        flow = round(flow,5)
        man_flow = round(man_flow,5)
        origin_flow = round(origin_flow,5)
        flow_diff = round(flow_diff,5)
        om_flow_diff = round(om_flow_diff,5)
                
        if row == 3 and cal_full_flag < 3:
            judge_index = 0
            for i in range(0,record_index):   
                curTm['time_diff'].append(CalTimeDiffus(realTm['time_diff'][judge_index])) 
               
                curTm['wave_time'].append(int(realTm['wave_time'][judge_index]))
                curTm['x_axis'].append(realTm['x_axis'][judge_index])
                
                hisTm['time_diff'].append(CalTimeDiffus(realTm['time_diff'][judge_index])) 
    
                hisTm['wave_time'].append(int(realTm['wave_time'][judge_index]))
                hisTm['x_axis'].append(realTm['x_axis'][judge_index])

                judge_index = next_index(judge_index)
            
        

        if row >= 3:
            flowdic['flow_x'].append(row)
            flowdic['flow_rate'].append(CalFlowm3h(aver_auto_flow))
            flowdic['man_flow_rate'].append(CalFlowm3h(aver_man_flow))
            flowdic['origin_flow_rate'].append(CalFlowm3h(aver_origin_flow))
            flowdic['flow'].append(flow)
            flowdic['man_flow'].append(man_flow)
            flowdic['origin_flow'].append(origin_flow)
            flowdic['flow_diff'].append(flow_diff)
            flowdic['om_flow_diff'].append(om_flow_diff)
            # frame_data_e.append([])
            # frame_data_e[-1] = [0 for i in range(72)]
            fv.AutoExcFlowData.append({})
            fv.AutoExcFlowData[-1] = copy.deepcopy(hisTm)
            fv.AutoExcFlowData[-1].update(copy.deepcopy(flowdic))

            fv.OriginFlowData.append({})
            fv.OriginFlowData[-1] = copy.deepcopy(curTm)
            
            # frame_data[row-3][1][0].extend(hisTm['x_axis'])
            # frame_data[row-3][1][1].extend(hisTm['wave_time'])
            # frame_data[row-3][1][2].extend(hisTm['time_diff'])

            hisTm.clear()
            hisTm =  {'wave_time' : [],'time_diff' : [],'x_axis':[]}
            curTm.clear()
            curTm =  {'wave_time' : [],'time_diff' : [],'man_time_diff' : [],'x_axis':[]}
            del(flowdic['flow_x'][0])
            del(flowdic['flow_rate'][0])
            del(flowdic['man_flow_rate'][0])
            del(flowdic['origin_flow_rate'][0])
            del(flowdic['flow'][0])
            del(flowdic['man_flow'][0])
            del(flowdic['flow_diff'][0])
            del(flowdic['origin_flow'][0])
            del(flowdic['om_flow_diff'][0])


             


        
        
        # total.cell(row+1, 16+1).value = sumflow
        # total.cell(row+1, 16+2).value = flowrate
        # total.cell(row+1, 16+3).value = cal_flag
        # total.cell(row+1, 16+4).value = select_time_diff_flag
        # total.cell(row+1, 16+5).value = select_wave_time_flag
        # total.cell(row+1, 16+6).value = select_time_diff_index
        # total.cell(row+1, 16+7).value = select_wave_time_index
    self.slider_rangeChanged_signal.emit(datanrows-4)
    



    # result_file_name = file_path.split('.', 1)
    # result_file_name = ''.join(result_file_name[0])
    # result_file_name = result_file_name + '_推导.xlsx'
    # result.save(result_file_name)



